"""
Secfiling_Extraction_Vivek - SEC Filing Extraction Tool
Automated extraction of oil & gas metrics from 10-Q and 10-K filings
with Excel export and PostgreSQL database storage
"""

import os
import sys
import json
import re
import time
import shutil
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Third-party imports
try:
    import requests
    from bs4 import BeautifulSoup
    from langchain_community.embeddings import HuggingFaceEmbeddings
    from langchain_openai import ChatOpenAI, AzureChatOpenAI
    from langchain_community.document_loaders import TextLoader
    from langchain_text_splitters import RecursiveCharacterTextSplitter
    from langchain_community.vectorstores import FAISS
    from langchain_core.prompts import ChatPromptTemplate
    from langchain_core.output_parsers import StrOutputParser
    import psycopg2
    print("✅ All imports successful")
except ImportError as e:
    print(f"❌ Import error: {e}")
    print("\n📦 Please install required packages:")
    print("pip install requests beautifulsoup4 langchain-community langchain-openai")
    print("pip install langchain-text-splitters faiss-cpu sentence-transformers openai")
    print("pip install pandas openpyxl psycopg2-binary")
    sys.exit(1)

# Environment setup
os.environ['TOKENIZERS_PARALLELISM'] = "False"

# Create directory structure
DIRS = {
    "extracted": Path("data/extracted"),
    "debug": Path("data/debug"),
    "output": Path("output")
}

for dir_path in DIRS.values():
    dir_path.mkdir(parents=True, exist_ok=True)

# Database configuration: Will prompt user or read environment variables at runtime

def prompt_db_credentials() -> Optional[Dict[str, object]]:
    """Prompt the user for PostgreSQL credentials, allowing skip.

    Returns a dict suitable for psycopg2.connect or None if skipped.
    """
    print("\n🔐 Database credentials not set. Press Enter to skip and only create Excel files.")
    # Try environment variables first
    env_host = os.environ.get('PGHOST') or os.environ.get('POSTGRES_HOST') or 'localhost'
    env_port = os.environ.get('PGPORT') or os.environ.get('POSTGRES_PORT') or '5432'
    env_db = os.environ.get('PGDATABASE') or os.environ.get('POSTGRES_DB')
    env_user = os.environ.get('PGUSER') or os.environ.get('POSTGRES_USER')
    env_password = os.environ.get('PGPASSWORD') or os.environ.get('POSTGRES_PASSWORD')

    try:
        host = input(f"Host [{env_host}]: ").strip() or env_host
        port_in = input(f"Port [{env_port}]: ").strip() or env_port
        database = input(f"Database{f' [{env_db}]' if env_db else ''}: ").strip() or (env_db or '')
        user = input(f"User{f' [{env_user}]' if env_user else ''}: ").strip() or (env_user or '')
        # Password: do not echo if possible
        try:
            import getpass
            password = getpass.getpass("Password (leave blank to use env or skip): ")
        except Exception:
            password = input("Password (leave blank to use env or skip): ")

        # Use env password if not provided explicitly
        if password == "" and env_password:
            password = env_password

        # If user chooses to skip (no database name or user or password), treat as skip
        if not database or not user or not password:
            print("➡️  Skipping database connection. Proceeding with Excel generation only.\n")
            return None

        try:
            port = int(port_in)
        except ValueError:
            port = 5432

        return {
            "host": host or "localhost",
            "port": port,
            "database": database,
            "user": user,
            "password": password
        }
    except KeyboardInterrupt:
        print("\n➡️  Skipping database connection. Proceeding with Excel generation only.\n")
        return None

# ============================================================================
# COMPANY DATABASE
# ============================================================================

COMPANY_DATABASE = {
    "FANG": {"name": "Diamondback Energy", "cik": "0001539838"},
    "PR": {"name": "Permian Resources", "cik": "0001658566"},
    "DVN": {"name": "Devon Energy", "cik": "0001090012"},
    "OVV": {"name": "Ovintiv", "cik": "0001792580"},
    "CTRA": {"name": "Coterra Energy", "cik": "0000858470"},
    "EOG": {"name": "EOG Resources", "cik": "0000821189"},
    "SM": {"name": "SM Energy", "cik": "0000893538"},
    "VTLE": {"name": "Vital Energy", "cik": "0001528129"},
    "MTDR": {"name": "Matador Resources", "cik": "0001520006"},
    "OXY": {"name": "Occidental Petroleum", "cik": "0000797468"},
    "XOM": {"name": "Exxon Mobil", "cik": "0000034088"}
}

# ============================================================================
# DATABASE FUNCTIONS
# ============================================================================

def get_db_connection(config: Dict[str, object]):
    """Create and return a database connection using provided config."""
    try:
        conn = psycopg2.connect(**config)
        return conn
    except psycopg2.Error as e:
        print(f"❌ Database connection failed: {e}")
        return None

def create_database_tables(conn):
    """Create all necessary database tables."""
    try:
        cursor = conn.cursor()
        
        # Table 1: Company Summary
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS company_summary (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                cik VARCHAR(20),
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                time_period TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 2: Production Data (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS production_data (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                time_period TEXT,
                quarter VARCHAR(10),
                year VARCHAR(10),
                oil_mbbl_per_day VARCHAR(50),
                ngl_mbbl_per_day VARCHAR(50),
                gas_mmcf_per_day VARCHAR(50),
                boe_mboe_per_day VARCHAR(50),
                oil_mmbls_total VARCHAR(50),
                ngl_mmbls_total VARCHAR(50),
                gas_bcf_total VARCHAR(50),
                boe_mmboe_total VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 3: Activity & Well Information (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS activity_wells (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                quarter VARCHAR(10),
                year VARCHAR(10),
                drilling_rigs VARCHAR(50),
                gross_wells_drilled VARCHAR(50),
                gross_wells_completed VARCHAR(50),
                gross_wells_til VARCHAR(50),
                net_wells_til VARCHAR(50),
                avg_lateral_length_drilled VARCHAR(50),
                avg_lateral_length_completed VARCHAR(50),
                working_interest_percent VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 4: Revenue Data (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS revenue_data (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                quarter VARCHAR(10),
                year VARCHAR(10),
                oil_revenue VARCHAR(50),
                ngl_revenue VARCHAR(50),
                gas_revenue VARCHAR(50),
                total_revenue VARCHAR(50),
                revenue_per_boe VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 5: Realized Pricing (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS realized_prices (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                quarter VARCHAR(10),
                year VARCHAR(10),
                oil_price VARCHAR(50),
                ngl_price VARCHAR(50),
                gas_price VARCHAR(50),
                boe_price VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 6: Cost Data (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS cost_data (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                filing_type VARCHAR(10),
                filing_date DATE,
                quarter VARCHAR(10),
                year VARCHAR(10),
                production_cost_per_boe VARCHAR(50),
                lease_operating_expense_per_boe VARCHAR(50),
                transportation_cost_per_boe VARCHAR(50),
                production_taxes_per_boe VARCHAR(50),
                development_capex VARCHAR(50),
                exploration_capex VARCHAR(50),
                total_capex VARCHAR(50),
                ddna_per_boe VARCHAR(50),
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, filing_date, filing_type)
            );
        """)
        
        # Table 7: Basin Data (Combined Value+Unit Storage)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS basin_data (
                id SERIAL PRIMARY KEY,
                ticker VARCHAR(10) NOT NULL,
                company_name VARCHAR(255),
                sec_filing_date DATE,
                file_type VARCHAR(10),
                basin_name VARCHAR(100),
                
                -- Gas Production (combined value+unit as text)
                gas_reserves VARCHAR(50),
                gas_per_day VARCHAR(50),
                
                -- Oil Production (combined value+unit as text)
                oil_reserves VARCHAR(50),
                oil_per_day VARCHAR(50),
                
                -- NGL Production (combined value+unit as text)
                ngl_reserves VARCHAR(50),
                ngl_per_day VARCHAR(50),
                
                -- Total BOE (combined value+unit as text)
                total_boe VARCHAR(50),
                boe_per_day VARCHAR(50),
                
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(ticker, sec_filing_date, file_type, basin_name)
            );
        """)
        
        conn.commit()
        cursor.close()
        print("✅ Database tables created successfully")
        return True
        
    except psycopg2.Error as e:
        print(f"❌ Error creating tables: {e}")
        conn.rollback()
        return False

# ============================================================================
# SEC EDGAR API FUNCTIONS (from main_extractor.py)
# ============================================================================

class SECFilingFetcher:
    """Handles fetching and parsing SEC filings"""
    
    BASE_URL = "https://www.sec.gov"
    HEADERS = {
        'User-Agent': 'Secfiling_Extraction_Vivek contact@example.com',
        'Accept-Encoding': 'gzip, deflate',
        'Host': 'www.sec.gov'
    }
    
    def __init__(self, cik: str, ticker: str):
        self.cik = cik.strip().zfill(10)
        self.ticker = ticker.upper()
        
    def get_filings_list(self, filing_type: str = "10-Q", count: int = 10) -> List[Dict]:
        """Fetch list of available filings from SEC EDGAR"""
        try:
            submissions_url = f"{self.BASE_URL}/cgi-bin/browse-edgar"
            params = {
                'action': 'getcompany',
                'CIK': self.cik,
                'type': filing_type,
                'dateb': '',
                'owner': 'exclude',
                'count': count,
                'search_text': ''
            }
            
            print(f"  → Fetching {filing_type} filings for CIK {self.cik}...")
            response = requests.get(submissions_url, headers=self.HEADERS, params=params, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            filings = []
            table = soup.find('table', {'class': 'tableFile2'})
            
            if not table:
                print("  ⚠ No filings table found")
                return []
            
            rows = table.find_all('tr')[1:]
            
            for row in rows:
                cols = row.find_all('td')
                if len(cols) >= 4:
                    filing_type_cell = cols[0].text.strip()
                    
                    if filing_type_cell == filing_type:
                        documents_link = cols[1].find('a')
                        filing_date = cols[3].text.strip()
                        
                        if documents_link:
                            href = documents_link['href']
                            accession = href.split('/')[-1]
                            
                            filings.append({
                                'type': filing_type_cell,
                                'date': filing_date,
                                'accession': accession,
                                'documents_url': f"{self.BASE_URL}{href}"
                            })
            
            print(f"  ✓ Found {len(filings)} {filing_type} filing(s)")
            return filings
            
        except Exception as e:
            print(f"  ✗ Error fetching filings: {e}")
            return []
    
    def get_filing_document_url(self, documents_url: str, filing_type: str) -> Optional[str]:
        """Get the primary document URL from the documents index page"""
        try:
            print(f"  → Fetching documents index page...")
            
            response = requests.get(documents_url, headers=self.HEADERS, timeout=30)
            response.raise_for_status()
            
            soup = BeautifulSoup(response.content, 'html.parser')
            doc_table = soup.find('table', {'class': 'tableFile'})
            
            if not doc_table:
                print("  ✗ Could not find document table")
                return None
            
            rows = doc_table.find_all('tr')
            
            primary_patterns = [r'\.htm$', r'\.html$']
            exclude_patterns = [
                r'index\.htm', r'_htm\.xml', r'\.xsd$', r'\.xml$',
                r'\.xls$', r'\.xlsx$', r'\.pdf$', r'\.jpg$',
                r'\.gif$', r'\.png$', r'ex\d+', r'graphic'
            ]
            
            candidate_docs = []
            
            for row in rows[1:]:
                cols = row.find_all('td')
                if len(cols) >= 3:
                    doc_link = cols[2].find('a')
                    if doc_link and doc_link.has_attr('href'):
                        doc_href = doc_link['href']
                        
                        is_primary = any(re.search(pattern, doc_href, re.IGNORECASE) for pattern in primary_patterns)
                        is_excluded = any(re.search(pattern, doc_href, re.IGNORECASE) for pattern in exclude_patterns)
                        
                        if is_primary and not is_excluded:
                            if doc_href.startswith('http'):
                                full_url = doc_href
                            elif doc_href.startswith('/'):
                                full_url = f"{self.BASE_URL}{doc_href}"
                            else:
                                base_path = '/'.join(documents_url.split('/')[:-1])
                                full_url = f"{base_path}/{doc_href}"
                            
                            priority = 0
                            if filing_type.lower().replace('-', '') in doc_href.lower():
                                priority = 3
                            elif self.ticker.lower() in doc_href.lower():
                                priority = 2
                            elif 'htm' in doc_href.lower() and 'xml' not in doc_href.lower():
                                priority = 1
                            
                            candidate_docs.append({
                                'url': full_url,
                                'href': doc_href,
                                'priority': priority
                            })
            
            if not candidate_docs:
                print("  ✗ No valid document links found")
                return None
            
            candidate_docs.sort(key=lambda x: x['priority'], reverse=True)
            best_doc = candidate_docs[0]
            print(f"  ✓ Found document: {best_doc['href']}")
            
            return best_doc['url']
            
        except Exception as e:
            print(f"  ✗ Error fetching document URL: {e}")
            return None

# ============================================================================
# TEXT EXTRACTION
# ============================================================================

def extract_text_from_url(url: str, output_path: Path) -> bool:
    """Extract text content from SEC filing URL"""
    try:
        headers = {
            'User-Agent': 'Secfiling_Extraction_Vivek contact@example.com',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
            'Accept-Encoding': 'gzip, deflate'
        }
        
        if '/ix?doc=' in url:
            import urllib.parse
            parsed = urllib.parse.urlparse(url)
            query_params = urllib.parse.parse_qs(parsed.query)
            if 'doc' in query_params:
                doc_path = query_params['doc'][0]
                url = f"https://www.sec.gov{doc_path}"
        
        print(f"  → Downloading filing from SEC...")
        response = requests.get(url, headers=headers, timeout=60)
        response.raise_for_status()
        
        print(f"  → Parsing HTML ({len(response.content):,} bytes)...")
        
        import warnings
        from bs4 import XMLParsedAsHTMLWarning
        warnings.filterwarnings("ignore", category=XMLParsedAsHTMLWarning)
        
        response_text = response.text
        response_text = re.sub(r'<ix:nonFraction[^>]*>(.*?)</ix:nonFraction>', r'\1', response_text, flags=re.DOTALL)
        response_text = re.sub(r'<ix:nonNumeric[^>]*>(.*?)</ix:nonNumeric>', r'\1', response_text, flags=re.DOTALL)
        try:
            soup = BeautifulSoup(response_text, 'lxml')
        except Exception:
            print("  ⚠ lxml parser not available, falling back to html.parser")
            soup = BeautifulSoup(response_text, 'html.parser')
        
        for element in soup(['script', 'style', 'meta', 'link', 'noscript', 'head']):
            element.decompose()
        
        body = soup.find('body')
        if body:
            text_content = body.get_text(separator='\n', strip=False)
        else:
            text_content = soup.get_text(separator='\n', strip=False)
        
        lines = []
        for line in text_content.splitlines():
            line = line.strip()
            if line and len(line) > 1 and not line.startswith('//'):
                lines.append(line)
        
        clean_text = '\n'.join(lines)
        clean_text = re.sub(r'\n{3,}', '\n\n', clean_text)
        clean_text = re.sub(r' {2,}', ' ', clean_text)
        
        if len(clean_text) < 1000:
            soup = BeautifulSoup(response.content, 'html.parser')
            for element in soup(['script', 'style', 'meta', 'link', 'noscript', 'head', 'header', 'footer', 'nav']):
                element.decompose()
            
            text_content = soup.get_text(separator=' ', strip=True)
            clean_text = re.sub(r'\s+', ' ', text_content)
            clean_text = clean_text.replace(' . ', '. ')
            clean_text = clean_text.replace(' , ', ', ')
            clean_text = re.sub(r'(\. )([A-Z])', r'.\n\2', clean_text)
        
        if len(clean_text) < 1000:
            print(f"  ✗ ERROR: Extraction failed - only {len(clean_text)} characters")
            return False
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(clean_text)
        
        print(f"  ✓ Extracted {len(clean_text):,} characters")
        
        # Save debug sample
        debug_path = DIRS['debug'] / f"{output_path.stem}_sample.txt"
        with open(debug_path, 'w', encoding='utf-8') as f:
            f.write(f"EXTRACTION VALIDATION SAMPLE\n{'='*80}\n")
            f.write(f"Total Length: {len(clean_text):,} characters\n")
            f.write(f"URL: {url}\n")
            f.write(f"{'='*80}\n\nFIRST 5000 CHARACTERS:\n{'-'*80}\n\n")
            f.write(clean_text[:5000])
        
        return True
        
    except Exception as e:
        print(f"  ✗ Extraction failed: {e}")
        return False

# ============================================================================
# RAG PIPELINE
# ============================================================================

def build_rag_pipeline(document_path: Path, embeddings_model, llm_model):
    """Build RAG pipeline for document analysis"""
    try:
        print(f"  → Loading document...")
        loader = TextLoader(str(document_path), encoding='utf-8')
        documents = loader.load()
        
        if not documents:
            print("  ✗ No documents loaded")
            return None
        
        print(f"  → Creating text chunks...")
        text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=8000,
            chunk_overlap=1500,
            length_function=len,
            separators=["\n\n", "\n", ". ", " ", ""]
        )
        
        splits = text_splitter.split_documents(documents)
        print(f"  ✓ Created {len(splits)} chunks")
        
        print(f"  → Building vector store...")
        vectorstore = FAISS.from_documents(splits, embeddings_model)
        retriever = vectorstore.as_retriever(search_type="similarity", search_kwargs={"k": 25})
        
        template = """
You are a professional financial data extraction specialist for SEC 10-Q and 10-K filings.

════════════════════════════════════════════════════════════════════════════════
CORE PRINCIPLES
════════════════════════════════════════════════════════════════════════════════
1. Extract ONLY factual numeric data explicitly stated in the context
2. NEVER infer, calculate, or approximate values
3. Combine value and unit into a SINGLE STRING (e.g., "148.4 MBbl/d", "$1,234 million")
4. NO separate unit or source fields
5. Use "Not found" for missing data points
6. Return strictly valid JSON format with simplified structure

════════════════════════════════════════════════════════════════════════════════
DATA STRUCTURE GUIDELINES
════════════════════════════════════════════════════════════════════════════════

**COMPANY-LEVEL (TOTAL) DATA:**
- Extract company-wide totals for production, revenue, costs, and activity
- These represent consolidated/aggregate figures across all operations
- Place in main sections: "production", "revenue", "costs", "activity"

**BASIN/SEGMENT DATA:**
- Extract the SAME metrics as company-level, but broken down by basin/segment
- Maintain consistent structure for ALL basins
- Include units with each value
- Common basins: Delaware Basin, Permian Basin, Midland Basin, Anadarko Basin, 
  Eagle Ford, Rockies, Haynesville, Marcellus, Montney, Uinta, etc.

════════════════════════════════════════════════════════════════════════════════
STANDARD UNITS (Preserve these exactly as found)
════════════════════════════════════════════════════════════════════════════════
Production:
  - Oil: MBbl/d (thousands barrels per day), MMBbl (million barrels total)
  - NGL: MBbl/d, MMBbl
  - Gas: MMcf/d (million cubic feet per day), Bcf (billion cubic feet)
  - BOE: MBoe/d (thousands barrels of oil equivalent per day), MMBoe (million BOE)

Revenue:
  - Million USD for totals (e.g., "$1,234 million" or "1,234")
  - Per unit: $/Bbl (oil, NGL), $/Mcf (gas), $/BOE (total)

Costs:
  - Per BOE: $/BOE
  - Capital expenditures: Million USD

Activity:
  - Rigs: count (integer)
  - Wells: count (gross/net)
  - Lateral length: feet
  - Working interest: percent (%)

════════════════════════════════════════════════════════════════════════════════
OUTPUT JSON SCHEMA (SIMPLIFIED)
════════════════════════════════════════════════════════════════════════════════
{{
  "time_period": "Three months ended June 30, 2025",
  "quarter": "Q2",
  "year": "2025",
  
  "production": {{
    "oil_production_mbbl_per_day": "148.4 MBbl/d",
    "oil_production_mmbl_total": "26.9 MMBbl",
    "ngl_production_mbbl_per_day": "113.6 MBbl/d",
    "ngl_production_mmbl_total": "20.6 MMBbl",
    "gas_production_mmcf_per_day": "3021.1 MMcf/d",
    "gas_production_bcf_total": "546.8 Bcf",
    "total_boe_mboe_per_day": "765.4 MBoe/d",
    "total_boe_mmboe_total": "138.5 MMBoe"
  }},
  
  "activity": {{
    "drilling_rigs": "19 rigs",
    "gross_wells_drilled": "372 wells",
    "gross_wells_completed": "410 wells",
    "gross_wells_til": "358 wells",
    "net_wells_til": "336 wells",
    "avg_lateral_length_drilled": "10215 feet",
    "avg_lateral_length_completed": "13402 feet",
    "working_interest_percent": "80%"
  }},
  
  "revenue": {{
    "oil_revenue": "$1,774 million",
    "ngl_revenue": "$425 million",
    "gas_revenue": "$1,499 million",
    "total_revenue": "$3,869 million",
    "revenue_per_boe": "$39.61/BOE",
    "oil_price_realized": "$66.52/Bbl",
    "ngl_price_realized": "$20.66/Bbl",
    "gas_price_realized": "$2.74/Mcf",
    "boe_price_realized": "$39.61/BOE"
  }},
  
  "costs": {{
    "production_cost_per_boe": "$17.78/BOE",
    "lease_operating_expense_per_boe": "$2.76/BOE",
    "transportation_cost_per_boe": "$4.00/BOE",
    "production_taxes_per_boe": "$1.32/BOE",
    "development_capex": "$1,121 million",
    "exploration_capex": "$14 million",
    "total_capex": "$1,192 million",
    "ddna_per_boe": "$7.83/BOE"
  }},
  
  "basins": {{
    "Delaware Basin": {{
      "oil_production_mbbl_per_day": "216 MBbl/d",
      "oil_production_mmbl_total": "80 MMBbl",
      "ngl_production_mbbl_per_day": "118 MBbl/d",
      "ngl_production_mmbl_total": "45 MMBbl",
      "gas_production_mmcf_per_day": "744 MMcf/d",
      "gas_production_bcf_total": "268 Bcf",
      "total_boe_mboe_per_day": "458 MBoe/d",
      "total_boe_mmboe_total": "170 MMBoe"
    }},
    "Permian Basin": {{
      "oil_production_mbbl_per_day": "Not found",
      "oil_production_mmbl_total": "Not found",
      "ngl_production_mbbl_per_day": "Not found",
      "ngl_production_mmbl_total": "Not found",
      "gas_production_mmcf_per_day": "Not found",
      "gas_production_bcf_total": "Not found",
      "total_boe_mboe_per_day": "Not found",
      "total_boe_mmboe_total": "Not found"
    }}
  }}
}}

════════════════════════════════════════════════════════════════════════════════
EXTRACTION RULES (SIMPLIFIED FORMAT)
════════════════════════════════════════════════════════════════════════════════

1. VALUE + UNIT COMBINATION:
   - Combine numeric value WITH unit in a single string
   - Examples:
     * "148.4 MBbl/d" (production rate)
     * "$1,774 million" (revenue)
     * "$66.52/Bbl" (price)
     * "19 rigs" (count)
     * "80%" (percentage)
   - Keep formatting simple and readable
   - If value not found, use: "Not found"

2. NUMBER FORMATTING:
   - Use commas for thousands where appropriate: "$1,234 million"
   - Preserve decimals as shown: "148.4 MBbl/d"
   - Keep negative signs for losses: "-$50 million"
   - Remove unnecessary zeros: "100 MMBbl" not "100.0 MMBbl"

3. UNIT STANDARDIZATION:
   - Use standard industry units:
     * Production: MBbl/d, MMBbl, MMcf/d, Bcf, MBoe/d, MMBoe
     * Revenue: $X million (always include $ and "million")
     * Prices: $/Bbl, $/Mcf, $/BOE
     * Costs: $/BOE, $X million
     * Activity: X rigs, X wells, X feet, X%
   - Match the document's scale (don't convert million to thousand, etc.)

4. BASIN IDENTIFICATION:
   - Common basin names: Delaware, Permian, Midland, Anadarko, Eagle Ford, 
     Bakken, DJ Basin, Powder River, Haynesville, Marcellus, Uinta, Rockies,
     Gulf of Mexico, Montney, Williston
   - Include ALL basins mentioned with production data
   - Use exact basin names as in document
   - For missing basin metrics, use "Not found"

5. HANDLING MISSING DATA:
   - Simply use "Not found" for any missing metric
   - Do not create placeholder or estimated values
   - Do not calculate or infer values

6. SPECIAL CASES:
   - Revenue fields: Always prefix with $ (e.g., "$1,234 million")
   - Price fields: Always use format like "$66.52/Bbl"
   - Percentage fields: Always include % symbol (e.g., "80%")
   - Count fields: Include unit (e.g., "19 rigs", "372 wells")
   - Quarter: Use Q1, Q2, Q3, Q4 format

════════════════════════════════════════════════════════════════════════════════
CONTEXT
════════════════════════════════════════════════════════════════════════════════
{context}

════════════════════════════════════════════════════════════════════════════════
TASK
════════════════════════════════════════════════════════════════════════════════
Extract all oil & gas operational and financial metrics from the context above.

REQUIREMENTS:
1. Extract company-level TOTAL metrics in main sections
2. Extract basin/segment-level breakdown with same metric structure
3. Combine value and unit in a SINGLE string (e.g., "148.4 MBbl/d", "$1,234 million")
4. NO separate "unit" or "source" fields
5. Use "Not found" for missing data
6. Return valid JSON matching the simplified schema above

Begin extraction:
"""
        
        prompt = ChatPromptTemplate.from_template(template.replace("{", "{{").replace("}", "}}").replace("{{context}}", "{context}"))
        
        chain = (
            {"context": retriever, "question": lambda x: x}
            | prompt
            | llm_model
            | StrOutputParser()
        )
        
        print(f"  ✓ RAG pipeline ready")
        return chain
        
    except Exception as e:
        print(f"  ✗ Pipeline build failed: {e}")
        return None

def create_extraction_prompt(filing_type: str) -> str:
    """Create extraction prompt"""
    return f"""Extract ALL oil & gas metrics from this {filing_type} filing.

REQUIRED METRICS:
- Production data (oil, NGL, gas, BOE)
- Activity/Well information
- Revenue metrics
- Cost metrics
- Basin/Segment breakdown

Return as structured JSON with all numeric values."""

def extract_metrics(rag_chain, filing_type: str) -> Dict:
    """Extract metrics using RAG pipeline"""
    try:
        print(f"  → Extracting metrics from {filing_type}...")
        
        prompt = create_extraction_prompt(filing_type)
        
        start_time = time.time()
        result = rag_chain.invoke(prompt)
        elapsed = time.time() - start_time
        
        print(f"  ✓ Extraction complete ({elapsed:.1f}s)")
        
        try:
            json_match = re.search(r'```json\s*(.*?)\s*```', result, re.DOTALL)
            if json_match:
                result = json_match.group(1)
            
            parsed = json.loads(result)
            return {"success": True, "data": parsed}
            
        except json.JSONDecodeError:
            return {"success": True, "data": result, "format": "text"}
            
    except Exception as e:
        print(f"  ✗ Extraction failed: {e}")
        return {"success": False, "error": str(e)}

# ============================================================================
# JSON TO EXCEL/DB PROCESSING FUNCTIONS
# ============================================================================

def extract_numeric_value(value_str):
    """Extract numeric value from simplified format string (e.g., '148.4 MBbl/d' -> 148.4)"""
    if not isinstance(value_str, str):
        return value_str
    
    if value_str.lower() in ['not found', 'n/m', 'nm', 'n.m.', '', 'none']:
        return None
    
    # Remove currency symbols and extract number
    pattern = r'-?\d{1,3}(?:,\d{3})*(?:\.\d+)?'
    cleaned = value_str.replace('$', '').replace('%', '')
    match = re.search(pattern, cleaned)
    
    if match:
        numeric_str = match.group().replace(',', '')
        try:
            return float(numeric_str)
        except ValueError:
            return None
    
    return None

def parse_simplified_value(value):
    """
    Parse value from simplified JSON format.
    Returns tuple of (numeric_value, original_string)
    Example: "148.4 MBbl/d" -> (148.4, "148.4 MBbl/d")
    """
    if not value or value == "Not found":
        return (None, value)
    
    if isinstance(value, dict):
        # Old format compatibility
        v = value.get('value', 'Not found')
        if v == 'Not found':
            return (None, 'Not found')
        unit = value.get('unit', '')
        return (extract_numeric_value(v), f"{v} {unit}".strip())
    
    # New simplified format - value is already a string like "148.4 MBbl/d"
    return (extract_numeric_value(value), value)

def get_value_from_dict(data_dict, possible_keys):
    """Try multiple keys to find value - returns (numeric, string) tuple"""
    for key in possible_keys:
        if key in data_dict:
            return parse_simplified_value(data_dict[key])
    return (None, None)

def extract_production_data(filing_data):
    """Extract production metrics from simplified format"""
    if not filing_data or 'production' not in filing_data:
        return {}
    
    prod = filing_data['production']
    
    return {
        'oil_mbbl_per_day': get_value_from_dict(prod, ['oil_production_mbbl_per_day', 'oil_mbbl_per_day'])[0],
        'oil_mbbl_per_day_str': get_value_from_dict(prod, ['oil_production_mbbl_per_day', 'oil_mbbl_per_day'])[1],
        'ngl_mbbl_per_day': get_value_from_dict(prod, ['ngl_production_mbbl_per_day', 'ngl_mbbl_per_day'])[0],
        'ngl_mbbl_per_day_str': get_value_from_dict(prod, ['ngl_production_mbbl_per_day', 'ngl_mbbl_per_day'])[1],
        'gas_mmcf_per_day': get_value_from_dict(prod, ['gas_production_mmcf_per_day', 'gas_mmcf_per_day'])[0],
        'gas_mmcf_per_day_str': get_value_from_dict(prod, ['gas_production_mmcf_per_day', 'gas_mmcf_per_day'])[1],
        'boe_mboe_per_day': get_value_from_dict(prod, ['total_boe_mboe_per_day', 'boe_mboe_per_day'])[0],
        'boe_mboe_per_day_str': get_value_from_dict(prod, ['total_boe_mboe_per_day', 'boe_mboe_per_day'])[1],
        'oil_mmbls_total': get_value_from_dict(prod, ['oil_production_mmbl_total', 'oil_mmbls'])[0],
        'oil_mmbls_total_str': get_value_from_dict(prod, ['oil_production_mmbl_total', 'oil_mmbls'])[1],
        'ngl_mmbls_total': get_value_from_dict(prod, ['ngl_production_mmbl_total', 'ngl_mmbls'])[0],
        'ngl_mmbls_total_str': get_value_from_dict(prod, ['ngl_production_mmbl_total', 'ngl_mmbls'])[1],
        'gas_bcf_total': get_value_from_dict(prod, ['gas_production_bcf_total', 'gas_bcf'])[0],
        'gas_bcf_total_str': get_value_from_dict(prod, ['gas_production_bcf_total', 'gas_bcf'])[1],
        'boe_mmboe_total': get_value_from_dict(prod, ['total_boe_mmboe_total', 'boe_mmboe'])[0],
        'boe_mmboe_total_str': get_value_from_dict(prod, ['total_boe_mmboe_total', 'boe_mmboe'])[1]
    }

def extract_activity_data(filing_data):
    """Extract activity metrics from simplified format"""
    if not filing_data or 'activity' not in filing_data:
        return {}
    
    activity = filing_data['activity']
    result = {}
    
    # Optimize: call parse_simplified_value once per field
    for field in ['drilling_rigs', 'gross_wells_drilled', 'gross_wells_completed', 
                  'gross_wells_til', 'net_wells_til', 'avg_lateral_length_drilled',
                  'avg_lateral_length_completed', 'working_interest_percent']:
        num_val, str_val = parse_simplified_value(activity.get(field, ''))
        result[field] = num_val
        result[f'{field}_str'] = str_val
    
    return result

def extract_revenue_data(filing_data):
    """Extract revenue metrics from simplified format"""
    if not filing_data or 'revenue' not in filing_data:
        return {}
    
    revenue = filing_data['revenue']
    
    return {
        'oil_revenue_million': get_value_from_dict(revenue, ['oil_revenue_million_usd', 'oil_revenue'])[0],
        'oil_revenue_million_str': get_value_from_dict(revenue, ['oil_revenue_million_usd', 'oil_revenue'])[1],
        'ngl_revenue_million': get_value_from_dict(revenue, ['ngl_revenue_million_usd', 'ngl_revenue'])[0],
        'ngl_revenue_million_str': get_value_from_dict(revenue, ['ngl_revenue_million_usd', 'ngl_revenue'])[1],
        'gas_revenue_million': get_value_from_dict(revenue, ['gas_revenue_million_usd', 'gas_revenue'])[0],
        'gas_revenue_million_str': get_value_from_dict(revenue, ['gas_revenue_million_usd', 'gas_revenue'])[1],
        'total_revenue_million': get_value_from_dict(revenue, ['total_revenue_million_usd', 'total_revenue'])[0],
        'total_revenue_million_str': get_value_from_dict(revenue, ['total_revenue_million_usd', 'total_revenue'])[1],
        'revenue_per_boe': get_value_from_dict(revenue, ['revenue_per_boe_usd', 'revenue_per_boe'])[0],
        'revenue_per_boe_str': get_value_from_dict(revenue, ['revenue_per_boe_usd', 'revenue_per_boe'])[1]
    }

def extract_pricing_data(filing_data):
    """Extract pricing metrics from simplified format"""
    if not filing_data or 'revenue' not in filing_data:
        return {}
    
    revenue = filing_data['revenue']
    
    return {
        'realized_price_oil_per_bbl': get_value_from_dict(revenue, ['realized_price_oil_usd_per_bbl', 'oil_price_realized', 'oil_price_per_bbl'])[0],
        'realized_price_oil_per_bbl_str': get_value_from_dict(revenue, ['realized_price_oil_usd_per_bbl', 'oil_price_realized', 'oil_price_per_bbl'])[1],
        'realized_price_ngl_per_bbl': get_value_from_dict(revenue, ['realized_price_ngl_usd_per_bbl', 'ngl_price_realized', 'ngl_price_per_bbl'])[0],
        'realized_price_ngl_per_bbl_str': get_value_from_dict(revenue, ['realized_price_ngl_usd_per_bbl', 'ngl_price_realized', 'ngl_price_per_bbl'])[1],
        'realized_price_gas_per_mcf': get_value_from_dict(revenue, ['realized_price_gas_usd_per_mcf', 'gas_price_realized', 'gas_price_per_mcf'])[0],
        'realized_price_gas_per_mcf_str': get_value_from_dict(revenue, ['realized_price_gas_usd_per_mcf', 'gas_price_realized', 'gas_price_per_mcf'])[1],
        'realized_price_boe': get_value_from_dict(revenue, ['realized_price_boe_usd_per_boe', 'boe_price_realized', 'boe_price'])[0],
        'realized_price_boe_str': get_value_from_dict(revenue, ['realized_price_boe_usd_per_boe', 'boe_price_realized', 'boe_price'])[1]
    }

def extract_cost_data(filing_data):
    """Extract cost metrics from simplified format"""
    if not filing_data or 'costs' not in filing_data:
        return {}
    
    costs = filing_data['costs']
    result = {}
    
    # Optimize: call parse_simplified_value once per field
    cost_fields = [
        ('production_cost_per_boe', 'production_cost_per_boe'),
        ('lease_operating_expense_per_boe', 'lease_operating_expense_per_boe'),
        ('transportation_cost_per_boe', 'transportation_cost_per_boe'),
        ('production_taxes_per_boe', 'production_taxes_per_boe'),
        ('ddna_per_boe', 'ddna_per_boe')
    ]
    
    for result_key, json_key in cost_fields:
        num_val, str_val = parse_simplified_value(costs.get(json_key, ''))
        result[result_key] = num_val
        result[f'{result_key}_str'] = str_val
    
    # Handle capex fields with fallback
    capex_value = costs.get('development_capex_million_usd', costs.get('development_capex', ''))
    num_val, str_val = parse_simplified_value(capex_value)
    result['development_capex_million'] = num_val
    result['development_capex_million_str'] = str_val
    
    capex_value = costs.get('exploration_capex_million_usd', costs.get('exploration_capex', ''))
    num_val, str_val = parse_simplified_value(capex_value)
    result['exploration_capex_million'] = num_val
    result['exploration_capex_million_str'] = str_val
    
    capex_value = costs.get('total_capex_million_usd', costs.get('total_capex', ''))
    num_val, str_val = parse_simplified_value(capex_value)
    result['total_capex_million'] = num_val
    result['total_capex_million_str'] = str_val
    
    return result

def parse_json_file(json_path):
    """Parse JSON file and extract all data (updated for simplified JSON structure)"""
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        # Handle new simplified JSON structure
        ticker = data.get('companyName', 'UNKNOWN')
        cik = data.get('cik', '')
        company_name = data.get('companyFullName', '')
        filing_type = data.get('fileType')
        filing_date = data.get('secFilingDate')
        
        filing_data = data.get('data', {})
        if not filing_data:
            return None
        
        time_period = filing_data.get('time_period', '')
        quarter = filing_data.get('quarter', '')
        year = filing_data.get('year', '')
        
        result = {
            'company_info': {
                'ticker': ticker,
                'cik': cik,
                'company_name': company_name,
                'filing_type': filing_type,
                'filing_date': filing_date,
                'time_period': time_period,
                'quarter': quarter,
                'year': year
            },
            'production': extract_production_data(filing_data),
            'activity': extract_activity_data(filing_data),
            'revenue': extract_revenue_data(filing_data),
            'pricing': extract_pricing_data(filing_data),
            'costs': extract_cost_data(filing_data),
            'basins': [],  # Will be populated below
            'raw_filing_data': filing_data  # Keep full data for Excel basin processing
        }
        
        # Extract basin data from simplified format
        basins_dict = filing_data.get('basins', {})
        result['basins'] = basins_dict
        
        return result
        
    except Exception as e:
        print(f"Error parsing {json_path}: {e}")
        import traceback
        traceback.print_exc()
        return None

# ============================================================================
# EXCEL CREATION
# ============================================================================

def create_excel_workbook(all_data, output_path):
    """Create Excel workbook with multiple sheets - using simplified format with value+unit strings"""
    
    writer = pd.ExcelWriter(output_path, engine='openpyxl')
    
    # Sheet 1: Production Data - Company Level
    production_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Time Period': data['company_info']['time_period'],
            'Quarter': data['company_info'].get('quarter', ''),
            'Year': data['company_info'].get('year', ''),
            'Oil Production (MBbl/d)': data['production'].get('oil_mbbl_per_day_str', ''),
            'Oil Production Total (MMBbl)': data['production'].get('oil_mmbls_total_str', ''),
            'NGL Production (MBbl/d)': data['production'].get('ngl_mbbl_per_day_str', ''),
            'NGL Production Total (MMBbl)': data['production'].get('ngl_mmbls_total_str', ''),
            'Gas Production (MMcf/d)': data['production'].get('gas_mmcf_per_day_str', ''),
            'Gas Production Total (Bcf)': data['production'].get('gas_bcf_total_str', ''),
            'Total BOE (MBoe/d)': data['production'].get('boe_mboe_per_day_str', ''),
            'Total BOE (MMBoe)': data['production'].get('boe_mmboe_total_str', '')
        }
        production_records.append(record)
    
    df_production = pd.DataFrame(production_records)
    df_production.to_excel(writer, sheet_name='Production Data', index=False)
    
    # Sheet 2: Activity & Wells
    activity_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Quarter': data['company_info'].get('quarter', ''),
            'Year': data['company_info'].get('year', ''),
            'Drilling Rigs': data['activity'].get('drilling_rigs_str', ''),
            'Gross Wells Drilled': data['activity'].get('gross_wells_drilled_str', ''),
            'Gross Wells Completed': data['activity'].get('gross_wells_completed_str', ''),
            'Gross Wells TIL': data['activity'].get('gross_wells_til_str', ''),
            'Net Wells TIL': data['activity'].get('net_wells_til_str', ''),
            'Avg Lateral Length Drilled': data['activity'].get('avg_lateral_length_drilled_str', ''),
            'Avg Lateral Length Completed': data['activity'].get('avg_lateral_length_completed_str', ''),
            'Working Interest': data['activity'].get('working_interest_percent_str', '')
        }
        activity_records.append(record)
    
    df_activity = pd.DataFrame(activity_records)
    df_activity.to_excel(writer, sheet_name='Activity & Wells', index=False)
    
    # Sheet 3: Revenue
    revenue_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Quarter': data['company_info'].get('quarter', ''),
            'Year': data['company_info'].get('year', ''),
            'Oil Revenue': data['revenue'].get('oil_revenue_million_str', ''),
            'NGL Revenue': data['revenue'].get('ngl_revenue_million_str', ''),
            'Gas Revenue': data['revenue'].get('gas_revenue_million_str', ''),
            'Total Revenue': data['revenue'].get('total_revenue_million_str', ''),
            'Revenue per BOE': data['revenue'].get('revenue_per_boe_str', '')
        }
        revenue_records.append(record)
    
    df_revenue = pd.DataFrame(revenue_records)
    df_revenue.to_excel(writer, sheet_name='Revenue', index=False)
    
    # Sheet 4: Realized Prices
    pricing_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Quarter': data['company_info'].get('quarter', ''),
            'Year': data['company_info'].get('year', ''),
            'Oil Price': data['pricing'].get('realized_price_oil_per_bbl_str', ''),
            'NGL Price': data['pricing'].get('realized_price_ngl_per_bbl_str', ''),
            'Gas Price': data['pricing'].get('realized_price_gas_per_mcf_str', ''),
            'BOE Price': data['pricing'].get('realized_price_boe_str', '')
        }
        pricing_records.append(record)
    
    df_pricing = pd.DataFrame(pricing_records)
    df_pricing.to_excel(writer, sheet_name='Realized Prices', index=False)
    
    # Sheet 5: Costs
    cost_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Quarter': data['company_info'].get('quarter', ''),
            'Year': data['company_info'].get('year', ''),
            'Production Cost per BOE': data['costs'].get('production_cost_per_boe_str', ''),
            'LOE per BOE': data['costs'].get('lease_operating_expense_per_boe_str', ''),
            'Transportation Cost per BOE': data['costs'].get('transportation_cost_per_boe_str', ''),
            'Production Taxes per BOE': data['costs'].get('production_taxes_per_boe_str', ''),
            'Development CapEx': data['costs'].get('development_capex_million_str', ''),
            'Exploration CapEx': data['costs'].get('exploration_capex_million_str', ''),
            'Total CapEx': data['costs'].get('total_capex_million_str', ''),
            'DD&A per BOE': data['costs'].get('ddna_per_boe_str', '')
        }
        cost_records.append(record)
    
    df_costs = pd.DataFrame(cost_records)
    df_costs.to_excel(writer, sheet_name='Costs', index=False)
    
    # Sheet 6: Basin Production Data (Detailed) - Using simplified format
    basin_records = []
    for data in all_data:
        basins_dict = data.get('basins', {})
        
        for basin_name, basin_data in basins_dict.items():
            if not isinstance(basin_data, dict):
                continue
            
            # In simplified format, basin data is directly stored as value+unit strings
            record = {
                'Ticker': data['company_info']['ticker'],
                'Company Name': data['company_info']['company_name'],
                'Filing Type': data['company_info']['filing_type'],
                'Filing Date': data['company_info']['filing_date'],
                'Time Period': data['company_info']['time_period'],
                'Quarter': data['company_info'].get('quarter', ''),
                'Year': data['company_info'].get('year', ''),
                'Basin Name': basin_name,
                'Oil Production (MBbl/d)': basin_data.get('oil_production_mbbl_per_day', 'Not found'),
                'Oil Production Total (MMBbl)': basin_data.get('oil_production_mmbl_total', 'Not found'),
                'NGL Production (MBbl/d)': basin_data.get('ngl_production_mbbl_per_day', 'Not found'),
                'NGL Production Total (MMBbl)': basin_data.get('ngl_production_mmbl_total', 'Not found'),
                'Gas Production (MMcf/d)': basin_data.get('gas_production_mmcf_per_day', 'Not found'),
                'Gas Production Total (Bcf)': basin_data.get('gas_production_bcf_total', 'Not found'),
                'Total BOE (MBoe/d)': basin_data.get('total_boe_mboe_per_day', 'Not found'),
                'Total BOE (MMBoe)': basin_data.get('total_boe_mmboe_total', 'Not found')
            }
            basin_records.append(record)
    
    if basin_records:
        df_basins = pd.DataFrame(basin_records)
        df_basins.to_excel(writer, sheet_name='Basin Production', index=False)
    
    # Sheet 7: Company Summary
    summary_records = []
    for data in all_data:
        record = {
            'Ticker': data['company_info']['ticker'],
            'CIK': data['company_info']['cik'],
            'Company Name': data['company_info']['company_name'],
            'Filing Type': data['company_info']['filing_type'],
            'Filing Date': data['company_info']['filing_date'],
            'Time Period': data['company_info']['time_period']
        }
        summary_records.append(record)
    
    df_summary = pd.DataFrame(summary_records)
    df_summary.to_excel(writer, sheet_name='Company Summary', index=False)
    
    writer.close()
    format_excel_workbook(output_path)
    
    print(f"✅ Excel file created: {output_path}")

def format_excel_workbook(file_path):
    """Apply formatting to Excel with proper column widths to prevent truncation"""
    wb = load_workbook(file_path)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(horizontal="left", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Auto-size columns with minimum width to prevent truncation
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if cell.value:
                        # Calculate length properly
                        cell_length = len(str(cell.value))
                        max_length = max(max_length, cell_length)
                        
                        # Apply cell alignment for data cells
                        if cell.row > 1:
                            cell.alignment = cell_alignment
                            cell.border = border
                except:
                    pass
            
            # Set column width with proper padding, minimum 15, maximum 60
            adjusted_width = min(max(max_length + 3, 15), 60)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Freeze top row for easy scrolling
        ws.freeze_panes = 'A2'
    
    wb.save(file_path)

# ============================================================================
# DATABASE INSERTION
# ============================================================================

def check_duplicate_filing(company_name: str, filing_type: str, filing_date: str, conn) -> bool:
    """
    Check if a filing already exists in the database.
    Returns True if duplicate exists, False otherwise.
    """
    try:
        cursor = conn.cursor()
        cursor.execute("""
            SELECT COUNT(*) FROM company_summary
            WHERE company_name = %s AND filing_type = %s AND filing_date = %s
        """, (company_name, filing_type, filing_date))
        
        count = cursor.fetchone()[0]
        cursor.close()
        
        return count > 0
    except psycopg2.Error as e:
        print(f"⚠️  Error checking for duplicate: {e}")
        return False

def insert_data_to_database(all_data, conn):
    """Insert parsed data into PostgreSQL database with duplicate check"""
    
    # Helper function to clean values for database insertion
    def clean_val(val):
        """Convert 'Not found' or empty values to NULL"""
        return None if not val or val == 'Not found' else val
    
    try:
        cursor = conn.cursor()
        inserted_count = 0
        skipped_count = 0
        
        for data in all_data:
            ticker = data['company_info']['ticker']
            company_name = data['company_info']['company_name']
            cik = data['company_info']['cik']
            filing_type = data['company_info']['filing_type']
            filing_date = data['company_info']['filing_date']
            time_period = data['company_info']['time_period']
            
            # Check for duplicate before insertion
            if check_duplicate_filing(company_name, filing_type, filing_date, conn):
                print(f"⚠️  Skipping duplicate: {company_name} - {filing_type} - {filing_date}")
                skipped_count += 1
                continue
            
            # Insert Company Summary
            cursor.execute("""
                INSERT INTO company_summary (ticker, cik, company_name, filing_type, filing_date, time_period)
                VALUES (%s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET cik = EXCLUDED.cik, company_name = EXCLUDED.company_name, time_period = EXCLUDED.time_period
            """, (ticker, cik, company_name, filing_type, filing_date, time_period))
            
            # Insert Production Data (Combined Value+Unit Strings)
            prod = data['production']
            quarter = data['company_info'].get('quarter', '')
            year = data['company_info'].get('year', '')
            
            cursor.execute("""
                INSERT INTO production_data 
                (ticker, company_name, filing_type, filing_date, time_period, quarter, year,
                oil_mbbl_per_day, ngl_mbbl_per_day, gas_mmcf_per_day, boe_mboe_per_day,
                oil_mmbls_total, ngl_mmbls_total, gas_bcf_total, boe_mmboe_total)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET oil_mbbl_per_day = EXCLUDED.oil_mbbl_per_day,
                    ngl_mbbl_per_day = EXCLUDED.ngl_mbbl_per_day,
                    gas_mmcf_per_day = EXCLUDED.gas_mmcf_per_day,
                    boe_mboe_per_day = EXCLUDED.boe_mboe_per_day,
                    oil_mmbls_total = EXCLUDED.oil_mmbls_total,
                    ngl_mmbls_total = EXCLUDED.ngl_mmbls_total,
                    gas_bcf_total = EXCLUDED.gas_bcf_total,
                    boe_mmboe_total = EXCLUDED.boe_mmboe_total
            """, (ticker, company_name, filing_type, filing_date, time_period, quarter, year,
                  clean_val(prod.get('oil_mbbl_per_day_str')), clean_val(prod.get('ngl_mbbl_per_day_str')),
                  clean_val(prod.get('gas_mmcf_per_day_str')), clean_val(prod.get('boe_mboe_per_day_str')),
                  clean_val(prod.get('oil_mmbls_total_str')), clean_val(prod.get('ngl_mmbls_total_str')),
                  clean_val(prod.get('gas_bcf_total_str')), clean_val(prod.get('boe_mmboe_total_str'))))
            
            # Insert Activity Data (Combined Value+Unit Strings)
            activity = data['activity']
            cursor.execute("""
                INSERT INTO activity_wells
                (ticker, company_name, filing_type, filing_date, quarter, year,
                drilling_rigs, gross_wells_drilled, gross_wells_completed, gross_wells_til, net_wells_til,
                avg_lateral_length_drilled, avg_lateral_length_completed, working_interest_percent)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET drilling_rigs = EXCLUDED.drilling_rigs,
                    gross_wells_drilled = EXCLUDED.gross_wells_drilled,
                    gross_wells_completed = EXCLUDED.gross_wells_completed,
                    gross_wells_til = EXCLUDED.gross_wells_til,
                    net_wells_til = EXCLUDED.net_wells_til,
                    avg_lateral_length_drilled = EXCLUDED.avg_lateral_length_drilled,
                    avg_lateral_length_completed = EXCLUDED.avg_lateral_length_completed,
                    working_interest_percent = EXCLUDED.working_interest_percent
            """, (ticker, company_name, filing_type, filing_date, quarter, year,
                  clean_val(activity.get('drilling_rigs_str')),
                  clean_val(activity.get('gross_wells_drilled_str')),
                  clean_val(activity.get('gross_wells_completed_str')),
                  clean_val(activity.get('gross_wells_til_str')),
                  clean_val(activity.get('net_wells_til_str')),
                  clean_val(activity.get('avg_lateral_length_drilled_str')),
                  clean_val(activity.get('avg_lateral_length_completed_str')),
                  clean_val(activity.get('working_interest_percent_str'))))
            
            # Insert Revenue Data (Combined Value+Unit Strings)
            revenue = data['revenue']
            cursor.execute("""
                INSERT INTO revenue_data
                (ticker, company_name, filing_type, filing_date, quarter, year,
                oil_revenue, ngl_revenue, gas_revenue, total_revenue, revenue_per_boe)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET oil_revenue = EXCLUDED.oil_revenue,
                    ngl_revenue = EXCLUDED.ngl_revenue,
                    gas_revenue = EXCLUDED.gas_revenue,
                    total_revenue = EXCLUDED.total_revenue,
                    revenue_per_boe = EXCLUDED.revenue_per_boe
            """, (ticker, company_name, filing_type, filing_date, quarter, year,
                  clean_val(revenue.get('oil_revenue_million_str')),
                  clean_val(revenue.get('ngl_revenue_million_str')),
                  clean_val(revenue.get('gas_revenue_million_str')),
                  clean_val(revenue.get('total_revenue_million_str')),
                  clean_val(revenue.get('revenue_per_boe_str'))))
            
            # Insert Pricing Data (Combined Value+Unit Strings)
            pricing = data['pricing']
            cursor.execute("""
                INSERT INTO realized_prices
                (ticker, company_name, filing_type, filing_date, quarter, year,
                oil_price, ngl_price, gas_price, boe_price)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET oil_price = EXCLUDED.oil_price,
                    ngl_price = EXCLUDED.ngl_price,
                    gas_price = EXCLUDED.gas_price,
                    boe_price = EXCLUDED.boe_price
            """, (ticker, company_name, filing_type, filing_date, quarter, year,
                  clean_val(pricing.get('realized_price_oil_per_bbl_str')),
                  clean_val(pricing.get('realized_price_ngl_per_bbl_str')),
                  clean_val(pricing.get('realized_price_gas_per_mcf_str')),
                  clean_val(pricing.get('realized_price_boe_str'))))
            
            # Insert Cost Data (Combined Value+Unit Strings)
            costs = data['costs']
            cursor.execute("""
                INSERT INTO cost_data
                (ticker, company_name, filing_type, filing_date, quarter, year,
                production_cost_per_boe, lease_operating_expense_per_boe, 
                transportation_cost_per_boe, production_taxes_per_boe,
                development_capex, exploration_capex, total_capex, ddna_per_boe)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (ticker, filing_date, filing_type) DO UPDATE
                SET production_cost_per_boe = EXCLUDED.production_cost_per_boe,
                    lease_operating_expense_per_boe = EXCLUDED.lease_operating_expense_per_boe,
                    transportation_cost_per_boe = EXCLUDED.transportation_cost_per_boe,
                    production_taxes_per_boe = EXCLUDED.production_taxes_per_boe,
                    development_capex = EXCLUDED.development_capex,
                    exploration_capex = EXCLUDED.exploration_capex,
                    total_capex = EXCLUDED.total_capex,
                    ddna_per_boe = EXCLUDED.ddna_per_boe
            """, (ticker, company_name, filing_type, filing_date, quarter, year,
                  clean_val(costs.get('production_cost_per_boe_str')),
                  clean_val(costs.get('lease_operating_expense_per_boe_str')),
                  clean_val(costs.get('transportation_cost_per_boe_str')),
                  clean_val(costs.get('production_taxes_per_boe_str')),
                  clean_val(costs.get('development_capex_million_str')),
                  clean_val(costs.get('exploration_capex_million_str')),
                  clean_val(costs.get('total_capex_million_str')),
                  clean_val(costs.get('ddna_per_boe_str'))))
            
            # Insert Basin Production Data - store combined value+unit strings directly
            basins_dict = data.get('basins', {})
            
            for basin_name, basin_data in basins_dict.items():
                if not isinstance(basin_data, dict):
                    continue
                
                # Get combined strings directly from JSON (no parsing needed!)
                gas_reserves = basin_data.get('gas_production_bcf_total', 'Not found')
                gas_per_day = basin_data.get('gas_production_mmcf_per_day', 'Not found')
                oil_reserves = basin_data.get('oil_production_mmbl_total', 'Not found')
                oil_per_day = basin_data.get('oil_production_mbbl_per_day', 'Not found')
                ngl_reserves = basin_data.get('ngl_production_mmbl_total', 'Not found')
                ngl_per_day = basin_data.get('ngl_production_mbbl_per_day', 'Not found')
                total_boe = basin_data.get('total_boe_mmboe_total', 'Not found')
                boe_per_day = basin_data.get('total_boe_mboe_per_day', 'Not found')
                
                cursor.execute("""
                    INSERT INTO basin_data
                    (ticker, company_name, sec_filing_date, file_type, basin_name,
                     gas_reserves, gas_per_day,
                     oil_reserves, oil_per_day,
                     ngl_reserves, ngl_per_day,
                     total_boe, boe_per_day)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ON CONFLICT (ticker, sec_filing_date, file_type, basin_name) DO UPDATE
                    SET gas_reserves = EXCLUDED.gas_reserves,
                        gas_per_day = EXCLUDED.gas_per_day,
                        oil_reserves = EXCLUDED.oil_reserves,
                        oil_per_day = EXCLUDED.oil_per_day,
                        ngl_reserves = EXCLUDED.ngl_reserves,
                        ngl_per_day = EXCLUDED.ngl_per_day,
                        total_boe = EXCLUDED.total_boe,
                        boe_per_day = EXCLUDED.boe_per_day
                """, (ticker, company_name, filing_date, filing_type, basin_name,
                      clean_val(gas_reserves), clean_val(gas_per_day),
                      clean_val(oil_reserves), clean_val(oil_per_day),
                      clean_val(ngl_reserves), clean_val(ngl_per_day),
                      clean_val(total_boe), clean_val(boe_per_day)))
            
            inserted_count += 1
        
        conn.commit()
        cursor.close()
        
        if inserted_count > 0:
            print(f"✅ Data inserted into database successfully ({inserted_count} records)")
        if skipped_count > 0:
            print(f"⚠️  Skipped {skipped_count} duplicate record(s)")
        
        return True
        
    except psycopg2.Error as e:
        print(f"❌ Database insertion error: {e}")
        conn.rollback()
        return False

# ============================================================================
# USER INTERFACE
# ============================================================================

def print_header():
    """Print application header"""
    print("\n" + "="*80)
    print("SECFILING_EXTRACTION_VIVEK - SEC FILING EXTRACTION TOOL")
    print("Automated Oil & Gas Metrics Extraction + Excel + PostgreSQL")
    print("="*80 + "\n")

def get_company_selection() -> Tuple[str, str, str]:
    """Interactive company selection"""
    print("📋 AVAILABLE COMPANIES:")
    print("-" * 80)
    
    companies_list = []
    for idx, (ticker, info) in enumerate(COMPANY_DATABASE.items(), 1):
        print(f"  {idx:2d}. {ticker:6s} - {info['name']:<30s} (CIK: {info['cik']})")
        companies_list.append((ticker, info['cik'], info['name']))
    
    print("-" * 80)
    
    while True:
        selection = input("\nEnter company number or ticker: ").strip().upper()
        
        if selection.isdigit():
            idx = int(selection)
            if 1 <= idx <= len(companies_list):
                ticker, cik, name = companies_list[idx - 1]
                print(f"\n✓ Selected: {name} ({ticker})\n")
                return ticker, cik, name
        
        if selection in COMPANY_DATABASE:
            info = COMPANY_DATABASE[selection]
            print(f"\n✓ Selected: {info['name']} ({selection})\n")
            return selection, info['cik'], info['name']
        
        print("❌ Invalid selection. Please try again.")

def select_filings(filings_10q: List[Dict], filings_10k: List[Dict]) -> Tuple[Optional[Dict], Optional[Dict]]:
    """Interactive filing selection"""
    selected_10q = None
    selected_10k = None
    
    if filings_10q:
        print("\n📄 AVAILABLE 10-Q FILINGS:")
        print("-" * 80)
        for idx, filing in enumerate(filings_10q, 1):
            print(f"  {idx}. {filing['date']}")
        print("-" * 80)
        
        choice = input("\nSelect 10-Q number (or Enter to skip): ").strip()
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(filings_10q):
                selected_10q = filings_10q[idx - 1]
                print(f"✓ Selected 10-Q from {selected_10q['date']}")
    
    if filings_10k:
        print("\n📄 AVAILABLE 10-K FILINGS:")
        print("-" * 80)
        for idx, filing in enumerate(filings_10k, 1):
            print(f"  {idx}. {filing['date']}")
        print("-" * 80)
        
        choice = input("\nSelect 10-K number (or Enter to skip): ").strip()
        if choice.isdigit():
            idx = int(choice)
            if 1 <= idx <= len(filings_10k):
                selected_10k = filings_10k[idx - 1]
                print(f"✓ Selected 10-K from {selected_10k['date']}")
    
    return selected_10q, selected_10k

# ============================================================================
# MAIN EXECUTION
# ============================================================================

def save_output_file(data: Dict, company_name: str, output_dir: Path):
    """Save results to separate JSON files for each filing type with simplified structure"""
    saved_files = []
    
    for filing_type, filing_info in data.get('filings', {}).items():
        result = filing_info.get('extraction_result', {})
        data_field = result.get('data')
        
        if isinstance(data_field, str) and result.get('format') == 'text':
            try:
                result['data'] = json.loads(data_field)
                del result['format']
            except json.JSONDecodeError:
                pass
        
        # Create separate JSON for this filing
        filing_date = filing_info.get('filing_date', 'unknown')
        # Clean company name for filename (remove spaces, special chars)
        clean_company_name = company_name.replace(' ', '').replace(',', '').replace('.', '')
        output_filename = f"{clean_company_name}_{filing_type}_{filing_date}.json"
        output_path = output_dir / output_filename
        
        ticker = data.get('company', {}).get('ticker', 'UNKNOWN')
        
        # Create individual filing JSON structure with NEW SIMPLIFIED FORMAT
        individual_filing_data = {
            "companyName": ticker,
            "companyFullName": company_name,
            "cik": data.get('company', {}).get('cik', ''),
            "fileType": filing_type,
            "secFilingDate": filing_date,
            "accessionNumber": filing_info.get('accession'),
            "documentUrl": filing_info.get('document_url'),
            "extractionDate": data.get('extraction_date'),
            "data": result.get('data', {}) if result.get('success') else {}
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(individual_filing_data, f, indent=2)
        
        print(f"✅ JSON saved: {output_path}")
        saved_files.append(output_path)
    
    return saved_files

def main():
    """Main execution function"""
    
    print_header()
    
    # Obtain database credentials (optional) and connect
    conn = None
    db_enabled = False
    # If environment variables provide full config, use them silently
    env_full = all([
        os.environ.get('PGHOST') or os.environ.get('POSTGRES_HOST'),
        os.environ.get('PGPORT') or os.environ.get('POSTGRES_PORT'),
        os.environ.get('PGDATABASE') or os.environ.get('POSTGRES_DB'),
        os.environ.get('PGUSER') or os.environ.get('POSTGRES_USER'),
        os.environ.get('PGPASSWORD') or os.environ.get('POSTGRES_PASSWORD')
    ])
    config = None
    if env_full:
        config = {
            "host": os.environ.get('PGHOST') or os.environ.get('POSTGRES_HOST'),
            "port": int(os.environ.get('PGPORT') or os.environ.get('POSTGRES_PORT') or 5432),
            "database": os.environ.get('PGDATABASE') or os.environ.get('POSTGRES_DB'),
            "user": os.environ.get('PGUSER') or os.environ.get('POSTGRES_USER'),
            "password": os.environ.get('PGPASSWORD') or os.environ.get('POSTGRES_PASSWORD')
        }
    else:
        config = prompt_db_credentials()

    if config:
        print("🔌 Testing database connection...")
        conn = get_db_connection(config)
        if conn:
            db_enabled = True
            print("✅ Database connection successful\n")
        else:
            print("⚠️  Proceeding without database. Excel files will still be generated.\n")
    
    # Create database tables
    if db_enabled:
        print("📊 Creating database tables...")
        if not create_database_tables(conn):
            print("❌ Failed to create database tables")
            conn.close()
            sys.exit(1)
    
    # Check OpenAI configuration
    use_azure = bool(os.getenv('AZURE_OPENAI_ENDPOINT'))
    
    if use_azure:
        if not all([os.getenv('AZURE_OPENAI_API_KEY'), 
                    os.getenv('AZURE_OPENAI_ENDPOINT'),
                    os.getenv('AZURE_OPENAI_DEPLOYMENT')]):
            print("❌ ERROR: Azure OpenAI environment variables not set")
            sys.exit(1)
    else:
        if not os.getenv('OPENAI_API_KEY'):
            print("❌ ERROR: OPENAI_API_KEY environment variable not set")
            print("\nPlease set your OpenAI API key:")
            print("  Windows: $env:OPENAI_API_KEY='your-key-here'")
            sys.exit(1)
    
    print("🤖 Initializing AI models...")
    try:
        if use_azure:
            llm = AzureChatOpenAI(
                api_key=os.getenv('AZURE_OPENAI_API_KEY'),
                azure_endpoint=os.getenv('AZURE_OPENAI_ENDPOINT'),
                azure_deployment=os.getenv('AZURE_OPENAI_DEPLOYMENT'),
                api_version=os.getenv('AZURE_OPENAI_API_VERSION', '2024-02-15-preview'),
                request_timeout=180
            )
            print(f"✓ Using Azure OpenAI")
        else:
            llm = ChatOpenAI(
                model="gpt-4o-mini",
                temperature=0,
                request_timeout=180
            )
            print("✓ Using OpenAI gpt-4o-mini")
        
        embeddings = HuggingFaceEmbeddings(
            model_name="sentence-transformers/all-MiniLM-L6-v2",
            model_kwargs={'device': 'cpu'}
        )
        print("✓ Models initialized successfully\n")
    except Exception as e:
        print(f"❌ Failed to initialize models: {e}")
        conn.close()
        sys.exit(1)
    
    # Company and filing selection
    ticker, cik, company_name = get_company_selection()
    
    print("🔍 Fetching available filings from SEC EDGAR...")
    fetcher = SECFilingFetcher(cik, ticker)
    
    filings_10q = fetcher.get_filings_list("10-Q", count=10)
    filings_10k = fetcher.get_filings_list("10-K", count=5)
    
    if not filings_10q and not filings_10k:
        print("\n❌ No filings found")
        if conn:
            conn.close()
        sys.exit(1)
    
    selected_10q, selected_10k = select_filings(filings_10q, filings_10k)
    
    if not selected_10q and not selected_10k:
        print("\n❌ No filings selected")
        if conn:
            conn.close()
        sys.exit(1)
    
    print("\n" + "="*80)
    print("PROCESSING SELECTED FILINGS")
    print("="*80 + "\n")
    
    results = {
        "company": {
            "ticker": ticker,
            "cik": cik,
            "name": company_name
        },
        "extraction_date": datetime.now().isoformat(),
        "filings": {}
    }
    
    filings_to_process = []
    if selected_10q:
        filings_to_process.append(("10-Q", selected_10q))
    if selected_10k:
        filings_to_process.append(("10-K", selected_10k))
    
    # Process each filing (check for duplicates first)
    for filing_type, filing in filings_to_process:
        print(f"\n{'='*80}")
        print(f"PROCESSING {filing_type} - Filed on {filing['date']}")
        print(f"{'='*80}\n")
        
        # Check if this filing already exists in database (only if DB enabled)
        if db_enabled and check_duplicate_filing(company_name, filing_type, filing['date'], conn):
            print(f"⚠️  Filing already exists in database: {company_name} - {filing_type} - {filing['date']}")
            print(f"⚠️  Skipping extraction for this filing.\n")
            continue
        
        filing_data = {
            "filing_date": filing['date'],
            "accession": filing['accession'],
            "document_url": None,
            "extraction_result": {"success": False, "error": "Not processed"}
        }
        
        doc_url = fetcher.get_filing_document_url(filing['documents_url'], filing_type)
        if not doc_url:
            print(f"❌ Could not get document URL")
            filing_data['extraction_result'] = {"success": False, "error": "Could not get document URL"}
            results['filings'][filing_type] = filing_data
            continue
        
        filing_data['document_url'] = doc_url
        
        print(f"\n[1/3] Extracting text from {filing_type}...")
        text_filename = f"{ticker}_{filing_type}_{filing['date']}.txt"
        text_path = DIRS['extracted'] / text_filename
        
        if not extract_text_from_url(doc_url, text_path):
            filing_data['extraction_result'] = {"success": False, "error": "Text extraction failed"}
            results['filings'][filing_type] = filing_data
            continue
        
        print(f"\n[2/3] Building RAG pipeline...")
        rag_chain = build_rag_pipeline(text_path, embeddings, llm)
        
        if not rag_chain:
            filing_data['extraction_result'] = {"success": False, "error": "RAG pipeline failed"}
            results['filings'][filing_type] = filing_data
            continue
        
        print(f"\n[3/3] Extracting metrics...")
        extraction_result = extract_metrics(rag_chain, filing_type)
        
        filing_data['extraction_result'] = extraction_result
        results['filings'][filing_type] = filing_data
    
    # Save JSON files (separate file per filing type)
    print("\n" + "="*80)
    print("SAVING RESULTS")
    print("="*80 + "\n")
    
    if not results['filings']:
        print("⚠️  No filings were processed (all may have been duplicates or failed)")
        conn.close()
        return
    
    # Save JSON alongside the script (current working directory)
    json_paths = save_output_file(results, company_name, Path.cwd())
    
    # Parse each JSON file and create Excel + insert to DB
    all_parsed_data = []
    excel_paths = []
    
    for json_path in json_paths:
        print(f"\n📊 Processing data from: {json_path.name}")
        parsed_data = parse_json_file(json_path)
        
        if parsed_data:
            all_parsed_data.append(parsed_data)
            
            # Create Excel file for this filing
            filing_type = parsed_data['company_info']['filing_type']
            filing_date = parsed_data['company_info']['filing_date']
            clean_company = company_name.replace(' ', '').replace(',', '').replace('.', '')
            excel_filename = f"{clean_company}_{filing_type}_{filing_date}.xlsx"
            # Save Excel to output folder
            excel_path = DIRS['output'] / excel_filename
            
            print(f"\n📈 Creating Excel workbook...")
            create_excel_workbook([parsed_data], excel_path)
            excel_paths.append(excel_path)
            
            # Insert into database if enabled
            if db_enabled:
                print(f"\n💾 Inserting data into PostgreSQL database...")
                if insert_data_to_database([parsed_data], conn):
                    print("✅ Data successfully stored in database")
                else:
                    print("⚠️  Excel created but database insertion failed")
        else:
            print(f"⚠️  Could not parse JSON file: {json_path.name}")
    
    if conn:
        conn.close()
    
    print("\n" + "="*80)
    print("EXTRACTION COMPLETE")
    print("="*80)
    print(f"\n📁 Output files:")
    for json_path in json_paths:
        print(f"   • JSON: {json_path}")
    for excel_path in excel_paths:
        print(f"   • Excel: {excel_path}")
    if db_enabled:
        print(f"   • Database: Connected and updated")
    else:
        print(f"   • Database: Skipped")
    print(f"\n   Total files processed: {len(all_parsed_data)}")
    print("\n" + "="*80 + "\n")
    print("✅ Excel files have been successfully saved in the 'output' folder.")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n❌ Process interrupted by user")
        sys.exit(0)
    except Exception as e:
        print(f"\n\n❌ Unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

